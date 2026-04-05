"""
Product / service hierarchy from Excel (Product List) for DC availability UI.

Reads ``data/product_catalog.xlsx`` sheet ``Ana Servis Kategorileri`` and maps rows
to AuraNotify ``datacenter-services`` category names.
"""

from __future__ import annotations

import logging
import re
import unicodedata
from collections import OrderedDict
from pathlib import Path
from typing import Any, Optional

logger = logging.getLogger(__name__)

_REPO_ROOT = Path(__file__).resolve().parents[2]
DEFAULT_CATALOG_PATH = _REPO_ROOT / "data" / "product_catalog.xlsx"

# (normalized substring in Excel service name, substring to find in normalized Aura category)
_SERVICE_HINTS: list[tuple[str, str]] = [
    ("yedek", "backup"),
    ("netbackup", "backup"),
    ("veeam", "backup"),
    ("zerto", "zerto"),
    ("replikasyon", "replication"),
    ("waf", "load balancer"),
    ("load balanc", "load balancer"),
    ("citrix", "load balancer"),
    ("ddos", "internet"),
    ("internet outage", "internet"),
    ("ipv4", "internet"),
    ("ipv6", "internet"),
    ("cross connection", "switch"),
    ("data switch", "switch"),
    ("management switch", "switch"),
    ("cloud dns", "internet"),
    ("s3 object", "storage"),
    ("object storage", "storage"),
    ("monitoring", "monitoring"),
    ("barindirma", "cabinet"),
    ("kabinet", "cabinet"),
    ("enerji birim", "cabinet power"),
    ("elektrik altyapi", "elektrik"),
    ("ups", "elektrik"),
    ("openstack", "openstack"),
    ("sap intel hana", "hyperconverged"),
    ("sap power hana", "hypervisor"),
    ("klasik mimari intel", "classic"),
    ("hyperconverged mimari intel vm", "hyperconverged mimari intel vm"),
    ("hyperconverged mimari intel dr", "hyperconverged"),
    ("klasik mimari intel dr", "classic"),
    ("fiziksel dedike firewall", "firewall"),
    ("sanal dedike firewall", "firewall"),
]

_cached_key: str | None = None
_cached_entries: list[dict[str, str]] = []


def _normalize(text: str) -> str:
    """Lowercase, strip, collapse spaces, strip accents for fuzzy match."""
    if not text:
        return ""
    t = unicodedata.normalize("NFKD", str(text))
    t = "".join(c for c in t if not unicodedata.combining(c))
    t = t.lower().strip()
    t = re.sub(r"\s+", " ", t)
    return t


def clear_service_hierarchy_cache() -> None:
    """Reset in-memory catalog (for tests)."""
    global _cached_key, _cached_entries
    _cached_key = None
    _cached_entries = []


def _load_excel_hierarchy(path: Path) -> list[dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        logger.warning("openpyxl not installed; service catalog unavailable")
        return []

    if not path.is_file():
        logger.warning("Product catalog Excel not found: %s", path)
        return []

    entries: list[dict[str, str]] = []
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        if "Ana Servis Kategorileri" not in wb.sheetnames:
            logger.warning("Sheet 'Ana Servis Kategorileri' missing in %s", path)
            return []
        ws = wb["Ana Servis Kategorileri"]
        rows = list(ws.iter_rows(values_only=True))
    except Exception as exc:
        logger.warning("Failed to read product catalog %s: %s", path, exc)
        return []

    if not rows:
        return []

    current_main = ""
    current_sub = ""
    for i, row in enumerate(rows):
        if i == 0:
            continue  # header
        if not row or len(row) < 3:
            continue
        raw_main, raw_sub, raw_svc = row[0], row[1], row[2]
        if raw_main is not None and str(raw_main).strip():
            current_main = str(raw_main).strip()
        if raw_sub is not None and str(raw_sub).strip():
            current_sub = str(raw_sub).strip()
        if raw_svc is None:
            continue
        svc = str(raw_svc).strip()
        if not svc or not current_main or not current_sub:
            continue
        entries.append(
            {
                "main_category": current_main,
                "sub_category": current_sub,
                "service": svc,
            }
        )
    return entries


def load_service_hierarchy(catalog_path: Path | None = None) -> list[dict[str, str]]:
    """Load flattened hierarchy rows from Excel (cached per path)."""
    global _cached_key, _cached_entries
    path = catalog_path or DEFAULT_CATALOG_PATH
    key = str(path.resolve())
    if _cached_key == key and _cached_entries:
        return list(_cached_entries)
    entries = _load_excel_hierarchy(path)
    _cached_key = key
    _cached_entries = entries
    return list(entries)


def nest_service_catalog(entries: list[dict[str, str]]) -> "OrderedDict[str, OrderedDict[str, list[str]]]":
    """main_category -> sub_category -> ordered list of service names."""
    tree: OrderedDict[str, OrderedDict[str, list[str]]] = OrderedDict()
    for e in entries:
        m = e.get("main_category") or ""
        s = e.get("sub_category") or ""
        svc = e.get("service") or ""
        if not m or not s or not svc:
            continue
        if m not in tree:
            tree[m] = OrderedDict()
        if s not in tree[m]:
            tree[m][s] = []
        tree[m][s].append(svc)
    return tree


def match_service_to_category(
    service_name: str,
    categories: list[dict[str, Any]],
) -> Optional[dict[str, Any]]:
    """
    Pick the AuraNotify category dict that best matches an Excel service name.

    Priority: exact normalized match > substring (min length 6) > keyword hints.
    """
    cats = [c for c in (categories or []) if isinstance(c, dict) and (c.get("category") or "").strip()]
    if not cats:
        return None
    ns = _normalize(service_name)
    if not ns:
        return None

    for c in cats:
        nc = _normalize(str(c.get("category")))
        if nc and ns == nc:
            return c

    best: Optional[dict[str, Any]] = None
    best_score = 0
    min_len = 6
    for c in cats:
        nc = _normalize(str(c.get("category")))
        if not nc:
            continue
        if len(nc) < min_len and len(ns) < min_len:
            continue
        if nc in ns or ns in nc:
            score = min(len(nc), len(ns))
            if score > best_score:
                best_score = score
                best = c
    if best is not None:
        return best

    for kw, aura_fragment in _SERVICE_HINTS:
        if kw not in ns:
            continue
        for c in cats:
            ncat = _normalize(str(c.get("category")))
            if aura_fragment in ncat:
                return c
    return None


def service_availability_pct(service_name: str, categories: list[dict[str, Any]]) -> tuple[float, Optional[dict[str, Any]]]:
    """
    Return (availability_pct, matched_category_or_none).
    When no AuraNotify category matches, availability is 100.0.
    """
    m = match_service_to_category(service_name, categories)
    if not m:
        return 100.0, None
    try:
        pct = float(m.get("availability_pct") or 100.0)
    except (TypeError, ValueError):
        pct = 100.0
    return pct, m
